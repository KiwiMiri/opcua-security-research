#!/usr/bin/env python3
"""
Create Excel file with results summary
Requires openpyxl: pip install openpyxl
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import os
    
    wb = Workbook()
    
    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Summary"
    
    # Headers
    headers = ["implementation", "language", "version", "commit_hash_or_pkginfo", 
               "build_env", "server_running", "server_port", "evidence_pcap"]
    ws1.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data
    data = [
        ["python-opcua", "Python", "0.98.13", "PyPI package: opcua==0.98.13",
         "OS: Linux aarch64 (Ubuntu 24.04); Python 3.12.3; pip install", "Yes", "4850", "pcaps/python_attack.pcap"],
        ["open62541", "C", "1.5.0", "git-hash: (source not available, using compiled binary), build version: 1.5.0",
         "GCC 13.3.0; CMake 3.28.3", "Yes", "4840", "pcaps/open62541_attack.pcap"],
        ["Node.js opcua", "Node.js", "2.157.0", "npm package: node-opcua@2.157.0",
         "Node v18.19.1; npm (global install)", "Yes", "4841", "pcaps/node_attack.pcap"],
        ["FreeOpcUa", "Python", "0.90.6", "PyPI package: freeopcua==0.90.6",
         "OS: Linux aarch64; Python 3.12.3; pip install", "Yes", "4842", "pcaps/freeopcua_attack.pcap"],
        ["Eclipse Milo", "Java", "0.6.9(SDK)", "pom.xml SDK version: 0.6.9",
         "Java OpenJDK 17.0.16; Maven 3.8.7", "Yes", "4844", "pcaps/milo_normal.pcap"],
        ["S2OPC", "C", "2554226f9", "git commit: 2554226f982c35c1e437cb0387eb4f347ea17865 (2025-10-20)",
         "OS: Linux aarch64; GCC 13.3.0; CMake 3.28.3", "Yes", "4840", "pcaps/N/A (not captured yet)"]
    ]
    
    for row in data:
        ws1.append(row)
    
    # Auto-adjust column widths
    for column in ws1.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 2: Hashes (PCAP checksums)
    ws2 = wb.create_sheet("Evidence_Hashes")
    ws2.append(["PCAP File", "SHA256 Hash"])
    ws2[1][0].fill = header_fill
    ws2[1][0].font = header_font
    ws2[1][1].fill = header_fill
    ws2[1][1].font = header_font
    
    # Read from pcap_sha256sums.txt
    if os.path.exists("pcaps/pcap_sha256sums.txt"):
        with open("pcaps/pcap_sha256sums.txt") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and len(line) > 65:
                    parts = line.split()
                    if len(parts) >= 2:
                        ws2.append([parts[1], parts[0]])
    
    # Auto-adjust column widths for hashes
    for column in ws2.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 100)
        ws2.column_dimensions[column_letter].width = adjusted_width
    
    # Save
    wb.save("results.xlsx")
    print("✓ results.xlsx created successfully!")
    print("  - Sheet 1: Summary")
    print("  - Sheet 2: Evidence_Hashes")
    
except ImportError:
    print("Error: openpyxl not installed. Install with: pip install openpyxl")
except Exception as e:
    print(f"Error creating Excel file: {e}")
    print("\nCreating TSV format instead...")
    
    with open("results.tsv", "w") as f:
        f.write("implementation\tlanguage\tversion\tcommit_hash_or_pkginfo\tbuild_env\tserver_running\tserver_port\tevidence_pcap\n")
        for row in data:
            f.write("\t".join(row) + "\n")
    print("✓ results.tsv created instead")

